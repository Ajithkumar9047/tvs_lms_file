# excel.py
from openpyxl import Workbook
import openpyxl
import os
from datetime import datetime
from Config import *
current_date = datetime.now().strftime('%d/%m/%Y')

def create_excel(rows_query1, rows_query2,overflowQUERY):
    dicts_query1 = [{'dealer_id': row[0], 'dealer_name': row[1], 'branch_id': row[2], 'threshold': row[3], 'created_on': row[4], 'id': row[5], 'pushed_date': row[6]} for row in rows_query1]
    dicts_query2 = [{'id': row[0], 'DealerId': row[1], 'Count': row[2], 'bucket': row[3], 'CreatedOn': row[4], 'maxcount': row[5], 'branchId': row[6], 'TotalCount': row[7], 'CrmPushed': row[8], 'thresholdGetdate': row[9], 'OverFlowCount': row[10], 'ModelId': row[11], 'dealer_id': row[12], 'dealer_name': row[13], 'branch_id': row[14], 'threshold': row[15], 'created_on': row[16], 'id': row[17], 'pushed_date': row[18]} for row in rows_query2]
    dicts_query3 = [{'source': row[0], 'name': row[1], 'mobile': row[2], 'part_id': row[3], 'email': row[4], 'model_name': row[5], 'city': row[6], 'model_id': row[7], 'pincode': row[8], 'DEALER_ID': row[9], 'Dealer_Name': row[10], 'Dealer_Mobile': row[11], 'Dealer_Emailid': row[12], 'AREA': row[13], 'Dealer_PinCode': row[14], 'threshold': row[15]} for row in overflowQUERY]
    #venv_directory = os.path.dirname(os.path.abspath(__file__))
    _, excel_file_name = tempfile.mkstemp(suffix=".xlsx", prefix=f"TVS_LMS_{current_date.replace('/', '_')}")
    #excel_file_name = os.path.join(venv_directory, f"TVS_LMS_{current_date.replace('/', '_')}.xlsx")
    wb = Workbook()

    ws_query1 = wb.active
    ws_query1.title = "Threshold Master"
    ws_query2 = wb.create_sheet(title="Threshold Feed")
    ws_query3 = wb.create_sheet(title="Threshold_Overflow")

    header_query1 = ["dealer_id", "dealer_name", "branch_id", "threshold", "created_on", "id", "pushed_date"]
    header_query2 = ['id', 'DealerId', 'Count', 'bucket', 'CreatedOn', 'maxcount', 'branchId', 'TotalCount', 'CrmPushed', 'thresholdGetdate', 'OverFlowCount', 'ModelId', 'dealer_id', 'dealer_name', 'branch_id', 'threshold', 'created_on', 'id', 'pushed_date']
    header_query3 = ['source','name','mobile','part_id','email','model_name','city','model_id','pincode','DEALER_ID','Dealer_Name','Dealer_Mobile','Dealer_Emailid','AREA','Dealer_PinCode']
    ws_query1.append(header_query1)
    ws_query2.append(header_query2)
    ws_query3.append(header_query3)
    
    for row in dicts_query1:
        ws_query1.append([row[key] for key in header_query1])
    for row in dicts_query2:
        ws_query2.append([row[key] for key in header_query2])
    for row in dicts_query3:
        ws_query2.append([row[key] for key in header_query3])

    border_style = openpyxl.styles.Side(style='thin')
    border = openpyxl.styles.Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    for row in ws_query1.iter_rows():
        for cell in row:
            cell.border = border
    for row in ws_query2.iter_rows():
        for cell in row:
            cell.border = border
    for row in ws_query3.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(excel_file_name)
    
    print("Excel file saved:", excel_file_name)
    return excel_file_name